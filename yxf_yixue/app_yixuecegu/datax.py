import os
import openpyxl
import openpyxl.utils
from utils._excel2db import Excel2Db


# 生成数据库，只需要在最初执行一次
def gen_db():
    c = Excel2Db('app_yixuecegu.db')
    c.transform2db(os.path.join(os.path.dirname(os.path.abspath(__file__)),'origin_data'))



class WangqiLuyao:
    def __init__(self):
        pass

    def execute_xlsx1(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_1(realpath2)")
        # 每次循环对应原始表格和目标表格的一行
        for i, row in enumerate(ws_origin):
            rows = ws_origin.max_row
            # 第一行为标题，需跳过
            if i == 0:
                continue
            # 循环到最末6行退出，防止计算六爻时出现空值
            if rows - i <= 6:
                break
            # for循环计数从0开始，但表格行列数从1开始，要注意差别
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_2(ws_origin, i)")
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    def execute_xlsx2(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_3(ws_origin, realpath2)")
        for i, row in enumerate(ws_origin):
            # 表中有标题行，数据第1行无法取未来值，需跳过前2次循环
            if i <= 1:
                continue
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_4(ws_origin, i)")
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    @staticmethod
    def check(typ):
        if typ == "上证指数":
            return "shangzhengzhishu"
        elif typ == "深证成指":
            return "shenzhengchengzhi"
        elif typ == "沪深300":
            return "hushen300"
        else:
            raise ImportError("输入参数错误：typ=" + typ)

    def shangzhengzhishu_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['日期',
             '卦号',
             'X1', 'X2', 'X3', 'X4', 'X5', 'X6',
             'XR1', 'XR2', 'XR3', 'XR4', 'XR5', 'XR6',
             'X7', 'X8', 'X9', 'X10', 'X11', 'X12', 'X13', 'X14', 'X15', 'X16',
             'X17', 'X18', 'X19', 'X20', 'X21', 'X22', 'X23', 'X24', 'X25', 'X26',
             'F'
             ])
        return wb_work, ws_work

    def shangzhengzhishu_2(self, ws_origin, i):
        # 当期数据
        DATE = ws_origin[i+1][0].value
        F1 = ws_origin[i+1][9].value
        try:
            X11 = ws_origin[i + 1][9].value-0
            X12 = ws_origin[i + 2][9].value-0
            X13 = ws_origin[i + 3][9].value-0
            X14 = ws_origin[i + 4][9].value-0
            X15 = ws_origin[i + 5][9].value-0
            X16 = ws_origin[i + 6][9].value-0
            X17 = ws_origin[i + 7][9].value - 0
            X18 = ws_origin[i + 8][9].value - 0
            X19 = ws_origin[i + 9][9].value - 0
        except:  # 原表中有无效记录，转化为0，如此，所有的无效记录都转化为乾卦
            X11 = 0
            X12 = 0
            X13 = 0
            X14 = 0
            X15 = 0
            X16 = 0
            X17 = 0
            X18 = 0
            X19 = 0
        X17=X11+X12
        X18=X13+X14
        X19=X15+X16
        X110=X11+X12+X13
        X111=X14+X15+X16
        X112=X11*X11
        X113=X12*X12
        X114=X13*X13
        X115=X14*X14
        X116=X15*X15
        X117=X16*X16
        X118=X11*X11*X11
        X119=X12*X12*X12
        X120=X13*X13*X13
        X121=X14*X14*X14
        X122=X15*X15*X15
        X123=X16*X16*X16
        X124=X11+X12+X13+X14
        X125=X11+X12+X13+X14+X15
        X126=X11+X12+X13+X14+X15+X16
        # 六爻数值变换
        # 0和正变1，负变0
        X11g = 1 if X11 >= 0 else 0
        X12g = 1 if X12 >= 0 else 0
        X13g = 1 if X13 >= 0 else 0
        X14g = 1 if X14 >= 0 else 0
        X15g = 1 if X15 >= 0 else 0
        X16g = 1 if X16 >= 0 else 0
        # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
        guashu1 = X11g*1 + X12g*2 + X13g*4 + X14g*8 + X15g*16 + X16g*32
        # 读取卦表
        wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\common_data\基础表-六十四卦.xlsx'))
        ws_gua = wb_gua.active
        GH1 = None
        # 查询卦表，求取卦号和卦名
        for row in ws_gua:
            if row[4].value == guashu1:
                GH1 = row[0].value
        arr = [
            NUM,  # 1,A
            GH1,  # 2,B
            X11, X12, X13, X14, X15, X16,  # 3-8,C-H
            X17, X18, X19, X110, X111, X112, X113, X114, X115, X116,  # 9-18,I-R
            X117, X118, X119, X120, X121, X122, X123, X124, X125, X126,  # 19-28,S-AB
            F1  # 29,AC
            ]
        return arr

    def shangzhengzhishu_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.append('R1')
        arr.append('RR1')
        ws_work.append(arr)
        return wb_work, ws_work

    def shangzhengzhishu_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        R1 = ws_origin[i][2].value
        RR1 = 1 if R1 >= 0 else -1
        arr.append(R1)
        arr.append(RR1)
        return arr


    def shenzhengchengzhi_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['NUM',  # 1,A
             'GH1',  # 2,B
             'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
             'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
             'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
             'F1'  # 29,AC
             ])
        return wb_work, ws_work


    def shenzhengchengzhi_2(self, ws_origin, i):
        # 当期数据
        NUM = ws_origin[i+1][0].value
        F1 = ws_origin[i+1][9].value
        try:
            X11 = ws_origin[i + 1][9].value - ws_origin[i + 2][9].value
            X12 = ws_origin[i + 2][9].value - ws_origin[i + 3][9].value
            X13 = ws_origin[i + 3][9].value - ws_origin[i + 4][9].value
            X14 = ws_origin[i + 4][9].value - ws_origin[i + 5][9].value
            X15 = ws_origin[i + 5][9].value - ws_origin[i + 6][9].value
            X16 = ws_origin[i + 6][9].value - ws_origin[i + 7][9].value
        except:  # 原表中有无效记录，转化为0
            X11 = 0
            X12 = 0
            X13 = 0
            X14 = 0
            X15 = 0
            X16 = 0
        X17=X11+X12
        X18=X13+X14
        X19=X15+X16
        X110=X11+X12+X13
        X111=X14+X15+X16
        X112=X11*X11
        X113=X12*X12
        X114=X13*X13
        X115=X14*X14
        X116=X15*X15
        X117=X16*X16
        X118=X11*X11*X11
        X119=X12*X12*X12
        X120=X13*X13*X13
        X121=X14*X14*X14
        X122=X15*X15*X15
        X123=X16*X16*X16
        X124=X11+X12+X13+X14
        X125=X11+X12+X13+X14+X15
        X126=X11+X12+X13+X14+X15+X16
        # 六爻数值变换
        # 0和正变1，负变0
        X11g = 1 if X11 >= 0 else 0
        X12g = 1 if X12 >= 0 else 0
        X13g = 1 if X13 >= 0 else 0
        X14g = 1 if X14 >= 0 else 0
        X15g = 1 if X15 >= 0 else 0
        X16g = 1 if X16 >= 0 else 0
        # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
        guashu1 = X11g*1 + X12g*2 + X13g*4 + X14g*8 + X15g*16 + X16g*32
        # 读取卦表
        wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\common_data\基础表-六十四卦.xlsx'))
        ws_gua = wb_gua.active
        GH1, GH2, GH3 = None, None, None
        # 查询卦表，求取卦号和卦名
        for row in ws_gua:
            if row[4].value == guashu1:
                GH1 = row[0].value
        arr = [
            NUM,  # 1,A
            GH1,  # 2,B
            X11, X12, X13, X14, X15, X16,  # 3-8,C-H
            X17, X18, X19, X110, X111, X112, X113, X114, X115, X116,  # 9-18,I-R
            X117, X118, X119, X120, X121, X122, X123, X124, X125, X126,  # 19-28,S-AB
            F1  # 29,AC
            ]
        return arr


    def shenzhengchengzhi_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.append('R1')
        arr.append('XFR1')
        ws_work.append(arr)
        return wb_work, ws_work


    def shenzhengchengzhi_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        XFR1, R1 = int(ws_origin[i][9].value)-int(ws_origin[i+1][9].value), int(ws_origin[i][9].value)
        arr.append(R1)
        arr.append(XFR1)
        return arr


    def hushen300_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['NUM',  # 1,A
             'GH1',  # 2,B
             'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
             'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
             'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
             'F1'  # 29,AC
             ])
        return wb_work, ws_work


    def hushen300_2(self, ws_origin, i):
        # 当期数据
        NUM = ws_origin[i+1][0].value
        F1 = ws_origin[i+1][9].value
        try:
            X11 = ws_origin[i + 1][9].value-0
            X12 = ws_origin[i + 2][9].value-0
            X13 = ws_origin[i + 3][9].value-0
            X14 = ws_origin[i + 4][9].value-0
            X15 = ws_origin[i + 5][9].value-0
            X16 = ws_origin[i + 6][9].value-0
        except:  # 原表中有无效记录，转化为0，如此，所有的无效记录都转化为乾卦
            X11 = 0
            X12 = 0
            X13 = 0
            X14 = 0
            X15 = 0
            X16 = 0
        X17=X11+X12
        X18=X13+X14
        X19=X15+X16
        X110=X11+X12+X13
        X111=X14+X15+X16
        X112=X11*X11
        X113=X12*X12
        X114=X13*X13
        X115=X14*X14
        X116=X15*X15
        X117=X16*X16
        X118=X11*X11*X11
        X119=X12*X12*X12
        X120=X13*X13*X13
        X121=X14*X14*X14
        X122=X15*X15*X15
        X123=X16*X16*X16
        X124=X11+X12+X13+X14
        X125=X11+X12+X13+X14+X15
        X126=X11+X12+X13+X14+X15+X16
        # 六爻数值变换
        # 0和正变1，负变0
        X11g = 1 if X11 >= 0 else 0
        X12g = 1 if X12 >= 0 else 0
        X13g = 1 if X13 >= 0 else 0
        X14g = 1 if X14 >= 0 else 0
        X15g = 1 if X15 >= 0 else 0
        X16g = 1 if X16 >= 0 else 0
        # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
        guashu1 = X11g*1 + X12g*2 + X13g*4 + X14g*8 + X15g*16 + X16g*32
        # 读取卦表
        wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\common_data\基础表-六十四卦.xlsx'))
        ws_gua = wb_gua.active
        GH1 = None
        # 查询卦表，求取卦号和卦名
        for row in ws_gua:
            if row[4].value == guashu1:
                GH1 = row[0].value
        arr = [
            NUM,  # 1,A
            GH1,  # 2,B
            X11, X12, X13, X14, X15, X16,  # 3-8,C-H
            X17, X18, X19, X110, X111, X112, X113, X114, X115, X116,  # 9-18,I-R
            X117, X118, X119, X120, X121, X122, X123, X124, X125, X126,  # 19-28,S-AB
            F1  # 29,AC
            ]
        return arr


    def hushen300_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.append('R1')
        arr.append('TF1')
        ws_work.append(arr)
        return wb_work, ws_work


    def hushen300_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        R1 = ws_origin[i][2].value
        TF1 = 1 if R1 >=0 else -1
        arr.append(R1)
        arr.append(TF1)
        return arr

