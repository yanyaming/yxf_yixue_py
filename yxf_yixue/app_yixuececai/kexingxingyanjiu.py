"""
可行性研究：
    对多种预测方法进行可行性研究，中奖率过低，投资额过大放弃；定胆、杀号结合
    胆号为和概率，越加越优，但投资额变大
    杀号为积概率，越加越小，而且呈指数级缩小，概率急剧变差，应当慎重
    概率分析还要注意概率的稳定性，即区间概率，在一个特殊的小区间内，区间概率可能会变得很差，比如连续很多期开出偏离总体概率的奖号
    倍投的概率论基础：不中奖事件的积概率，不中奖的概率呈指数级下降，中奖概率上升。但要考虑运气，即区间中奖概率可能很差，倍投投资额变得很大，坚持不下去
    数术预测有数术的规律，不需要倍投
    
    组选：放弃组三，取组六出现率70%（理论72%）；
    假设选4投4注8元，则需要最低区间概率7%才能稳赢；
    假设选5投10注20元（或2胆拖3码），则需要最低区间概率20%才能稳赢；（主要，投资低则时间风险大）
    假设选6投20注40元（或2胆拖4码），则需要最低区间概率35%才能稳赢；（主要，投资高则资金风险大）
    假设选7投35注70元，则需要最低区间概率60%才能稳赢；
    同样的选号方法直接扩展到直选1变6，则盈利更大。
    
    直选：定数胆+定位胆+杀号，每次投注25注50元以下，需要最低区间概率5%；
    
"""


import os
import openpyxl
import openpyxl.utils
import datetime
import yixuececai.fenxi as fenxi
import yixuececai.dulicecai as dulicecai
import xiaochengtu.xiaochengtu_api
import liuyao.liuyao_api
import jinkoujue.jinkoujue_api
import bazi.bazi_api


class Xuanhaodingdanfangfa:
    def __init__(self):
        pass

    def dingdan(self, row, zuliu):
        # zuliu_copy = zuliu.copy()
        # zuliu.clear()
        #
        # # 【定：开奖号*0.618】概率20%上下，是标准概率，没用
        # try:
        #     dan1 = []
        #     dan1_number = int((row[1].value*100 + row[2].value*10 + row[3].value)*0.618)
        #     for i, j in enumerate(str(dan1_number)):
        #         if i < 1:
        #             dan1.append(int(j))
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             for l in dan1:
        #                 if l == j[k]:
        #                     if j not in zuliu:
        #                         zuliu.append(j)
        #     print(len(zuliu))
        # except Exception as e:
        #     raise e

        # 【定：开奖号*3.14】概率
        # try:
        #     dan2 = []
        #     dan2_number = int((row[1].value * 100 + row[2].value * 10 + row[3].value) * 3.14)
        #     for i, j in enumerate(str(dan2_number)):
        #         if i < 1:
        #             dan2.append(int(j))
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             for l in dan2:
        #                 if l == j[k]:
        #                     if j not in zuliu:
        #                         zuliu.append(j)
        #     print(len(zuliu))
        # except Exception as e:
        #     raise e

        # 【定：测试】
        # try:
        #     dan3 = []
        #     dan3_number = int((row[1].value*4 + row[2].value*9 + row[3].value*9 + 3)%10)
        #     for i, j in enumerate(str(dan3_number)):
        #         if i < 1:
        #             dan3.append(int(j))
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             for l in dan3:
        #                 if l == j[k]:
        #                     if j not in zuliu:
        #                         zuliu.append(j)
        #     print(len(zuliu))
        # except Exception as e:
        #     raise e
        pass


class Xuanhaoshahaofangfa:
    # 编程问题：对列表进行for循环，中间删除某些项会打乱索引，所以要用while循环或者建立copy
    # 资料中说的概率90%以上的多是夸张宣传，实际效果并不理想
    # 杀号方法小于90%不采用，几乎没有能用的
    def __init__(self):
        pass

    def shahao(self, row, All_kebian):
        All_duizhao = All_kebian.copy()  # zuliu_copy建立索引和查询，zuliu保存结果

        # 【杀上期大小组合组合，组】概率：随机概率
        # try:
        #     # 判断模式
        #     daxiaomoshi = []
        #     if row[2].value in [0, 1, 2, 3, 4]:
        #         daxiaomoshi.append(0)
        #     else:
        #         daxiaomoshi.append(1)
        #     if row[3].value in [0, 1, 2, 3, 4]:
        #         daxiaomoshi.append(0)
        #     else:
        #         daxiaomoshi.append(1)
        #     if row[4].value in [0, 1, 2, 3, 4]:
        #         daxiaomoshi.append(0)
        #     else:
        #         daxiaomoshi.append(1)
        #     for i in All_duizhao:
        #         daxiaomoshi1 = []
        #         if i[0] in [0, 1, 2, 3, 4]:
        #             daxiaomoshi1.append(0)
        #         else:
        #             daxiaomoshi1.append(1)
        #         if i[1] in [0, 1, 2, 3, 4]:
        #             daxiaomoshi1.append(0)
        #         else:
        #             daxiaomoshi1.append(1)
        #         if i[2] in [0, 1, 2, 3, 4]:
        #             daxiaomoshi1.append(0)
        #         else:
        #             daxiaomoshi1.append(1)
        #         if daxiaomoshi1 == daxiaomoshi:
        #             if i in All_kebian:
        #                 All_kebian.remove(i)
        # except Exception as e:
        #     raise e

        # 【杀上期奇偶组合组合，组】概率：
        # try:
        #     kuaduhe = (abs(row[2].value - row[3].value) + abs(row[3].value - row[4].value) + abs(row[2].value - row[4].value))%10
        #     for i in All_duizhao:
        #         if (abs(i[0] - i[1]) + abs(i[1] - i[2]) + abs(i[0] - i[2]))%10 == kuaduhe:
        #             if i in All_kebian:
        #                 All_kebian.remove(i)
        # except Exception as e:
        #     raise e

        # 【杀跨度和法，组】概率：随机概率，无用
        # 杀与上期跨度和相同的组
        # try:
        #     kuaduhe = (abs(row[2].value - row[3].value) + abs(row[3].value - row[4].value) + abs(row[2].value - row[4].value))%10
        #     for i in All_duizhao:
        #         if (abs(i[0] - i[1]) + abs(i[1] - i[2]) + abs(i[0] - i[2]))%10 == kuaduhe:
        #             if i in All_kebian:
        #                 All_kebian.remove(i)
        # except Exception as e:
        #     raise e

        # 【杀开奖号按位相加，组】概率：67%
        # 上期开奖号按位+0、+1、+2，+3，+4，+5，+6，+7，+8，+9
        # try:
        #     for j in range(0, 10):
        #         # 百位
        #         if (row[1].value + j) > 9:
        #             v1 = row[1].value + j - 10
        #         else:
        #             v1 = row[1].value + j
        #         # 十位
        #         if (row[2].value + j) > 9:
        #             v2 = row[2].value + j - 10
        #         else:
        #             v2 = row[2].value + j
        #         # 个位
        #         if (row[3].value + j) > 9:
        #             v3 = row[3].value + j - 10
        #         else:
        #             v3 = row[3].value + j
        #         if sorted([v1, v2, v3]) in zuliu:
        #             zuliu.remove(sorted([v1, v2, v3]))
        # except Exception as e:
        #     raise e

        # 【杀连号，组】概率：60%
        # 固定连号。有多种分类。
        # 排列顺连号012/123/234/345/456/567/678/789/890/901；024/135/246/357/468/579/680/791/802/913。
        # 排列逆连号098/987/876/765/654/543/432/321/210/109；086/975/864/753/642/531/420/319/208/197。
        # 组合的连号012/123/234/345/456/567/678/789/089/019；024/135/246/357/468/579/068/179/028/139。
        # try:
        #     for j in zuliu_copy:
        #         # 间隔1连号前8组
        #         if (j[2] - j[1]) == 1 and (j[1] - j[0]) == 1:
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         # 间隔2连号前6组
        #         elif (j[2] - j[1]) == 2 and (j[1] - j[0]) == 2:
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #     if [0, 8, 9] in zuliu:
        #         zuliu.remove([0, 8, 9])  # 直接杀掉特定的组
        #     if [0, 1, 9] in zuliu:
        #         zuliu.remove([0, 1, 9])
        #     if [0, 6, 8] in zuliu:
        #         zuliu.remove([0, 6, 8])
        #     if [1, 7, 9] in zuliu:
        #         zuliu.remove([1, 7, 9])
        #     if [0, 2, 8] in zuliu:
        #         zuliu.remove([0, 2, 8])
        #     if [1, 3, 9] in zuliu:
        #         zuliu.remove([1, 3, 9])
        # except Exception as e:
        #     raise e

        # 【杀相同和值，组】概率：48%
        # 上期三位和值、前一期三位和值，以及两两组合，共八组和值，杀掉的组更多
        # try:
        #     for j in zuliu_copy:
        #         if (j[0] + j[1] + j[2]) == (row[1].value + row[2].value + row[3].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[1].value + row[2].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[1].value + row[3].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[2].value + row[3].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[7].value + row[8].value + row[9].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[7].value + row[8].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[7].value + row[9].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        #         elif (j[0] + j[1] + j[2]) == (row[8].value + row[9].value):
        #             if j in zuliu:
        #                 zuliu.remove(j)
        # except Exception as e:
        #     raise e

        # 【杀和尾及其运算，4码】概率：25%（杀4码剩6码，直选投注理论概率12%，此法可能有效？待验证）
        # 试机号和尾+开奖号和尾，取和尾的和尾、和尾%9、和尾%7、和尾%5
        # try:
        #     hewei = (row[1].value + row[2].value + row[3].value) % 10 + (row[4].value + row[5].value + row[6].value) % 10
        #     for j in All_duizhao:
        #         for k in range(0, 3):
        #             if j[k] == (hewei % 9):
        #                 if j in All_kebian:
        #                     All_kebian.remove(j)
        #             elif j[k] == (hewei % 10):
        #                 if j in All_kebian:
        #                     All_kebian.remove(j)
        #             elif j[k] == (hewei % 7):
        #                 if j in All_kebian:
        #                     All_kebian.remove(j)
        #             elif j[k] == (hewei % 5):
        #                 if j in All_kebian:
        #                     All_kebian.remove(j)
        # except Exception as e:
        #     raise e

        # 【杀两期开奖号运算，2码】概率：36%
        # 两期开奖号按位相减的绝对值相加，取和尾；一期开奖号按位平方和，取尾。
        # try:
        #     he = abs(row[1].value - row[7].value) + abs(row[2].value - row[8].value) + abs(row[3].value - row[9].value)
        #     pingfanghe = row[1].value**2 + row[2].value**2 + row[3].value**2
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             if j[k] == (he % 10):
        #                 if j in zuliu:
        #                     zuliu.remove(j)
        #             elif j[k] == (pingfanghe % 10):
        #                 if j in zuliu:
        #                     zuliu.remove(j)
        #             else:
        #                 pass
        # except Exception as e:
        #     raise e

        # 【杀路数和，1码】概率：51%
        # try:
        #     lushuhe = 0
        #     for j in [row[1].value, row[2].value, row[3].value]:
        #         if j in [0, 3, 6, 9]:
        #             lushuhe += 0
        #         elif j in [1, 4, 7]:
        #             lushuhe += 1
        #         elif j in [2, 5, 8]:
        #             lushuhe += 2
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             if j[k] == lushuhe:
        #                 if j in zuliu:
        #                     zuliu.remove(j)
        #             else:
        #                 pass
        # except Exception as e:
        #     raise e

        # 【当期期数的后两位，2码】概率：36%
        # try:
        #     qishuhouliangwei = [int(str(row[0].value)[-1]), int(str(row[0].value)[-2])]
        #     print(qishuhouliangwei)
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             if j[k] in qishuhouliangwei:
        #                 if j in zuliu:
        #                     zuliu.remove(j)
        #             else:
        #                 pass
        # except Exception as e:
        #     raise e

        # 【前两期的十位数，2码】概率：36%
        # try:
        #     qishuhouliangwei = [row[8].value, row[11].value]
        #     print(qishuhouliangwei)
        #     for j in zuliu_copy:
        #         for k in range(0, 3):
        #             if j[k] in qishuhouliangwei:
        #                 if j in zuliu:
        #                     zuliu.remove(j)
        #             else:
        #                 pass
        # except Exception as e:
        #     raise e
        pass


class Shushufangfa:
    def __init__(self):
        pass

    def xiaochengtu_canwuyishufa(self, row, All_kebian):
        # 定数胆
        # 时间起卦，四正归藏总概率：10%（随机概率，看不出有效）
        dt = row[0].value + datetime.timedelta(hours=22)
        Prk_duizhao = All_kebian.copy()
        # 时间起卦
        a = xiaochengtu.xiaochengtu_api.Api()
        a.paipan(dt, lingdongshu=None, shuziqigua=None, guizangfangfa='四正')
        # 获取预测结果，共7个数，可能有重复
        shu7 = a.get_canwuyishufa()
        # 选取符合定数胆的奖号
        All_kebian.clear()
        for item in Prk_duizhao:  # 每一项
            for w in range(0, 3):  # 每一位
                if item[w] == shu7[0]:  # 如果符合定数胆号（暂选整体定数）
                    if item not in All_kebian:
                        All_kebian.append(item)  # 选定

    def xiaochengtu_dingweidanfa(self, row, All_kebian):
        # 定位胆
        # 时间起卦，四正归藏总概率：13%（随机概率，看不出有效）
        # 时间起卦，
        dt = row[0].value + datetime.timedelta(hours=22)
        All_duizhao = All_kebian.copy()
        # 时间起卦
        a = xiaochengtu.xiaochengtu_api.Api()
        a.paipan(dt, lingdongshu=None, shuziqigua=None, guizangfangfa='四正')
        shu, wei = a.get_dingweidanfa()
        # 转化位代码
        wei_num = []
        if wei[0:1] == '百':
            wei_num.append(0)
        if wei[0:1] == '十':
            wei_num.append(1)
        if wei[0:1] == '个':
            wei_num.append(2)
        if len(wei) > 1:
            if wei[1:2] == '百':
                wei_num.append(0)
            if wei[1:2] == '十':
                wei_num.append(1)
            if wei[1:2] == '个':
                wei_num.append(2)
        # 选取符合定位胆的奖号
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            for w in range(0, 3):  # 每一位
                if w in wei_num:  # 如果是定位胆所在位
                    if item[w] == shu:  # 如果符合定位胆号
                        if item not in All_kebian:
                            All_kebian.append(item)  # 选定

    def dulicecai_wuxingheyifa(self, row, All_kebian):
        # 定数胆（随机概率，看不出有效）
        dt = row[0].value + datetime.timedelta(hours=22)
        All_duizhao = All_kebian.copy()
        num = []
        num.append(row[2].value)
        num.append(row[3].value)
        num.append(row[4].value)
        num.append(row[5].value)
        num.append(row[6].value)
        num.append(row[7].value)
        # 调用方法
        a = dulicecai.Dulicecai()
        shu = a.wuxingheyifa(num)
        # 选取符合定数胆的奖号
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            for w in range(0, 3):  # 每一位
                if item[w] == shu[0]:  # 如果符合定数胆号
                    if item not in All_kebian:
                        All_kebian.append(item)  # 选定

    def dulicecai_hezhiwuxingshahaofa(self, row, All_kebian):
        # 杀号。概率：50%（还是随机概率，无用）
        dt = row[0].value + datetime.timedelta(hours=22)
        All_duizhao = All_kebian.copy()
        a = dulicecai.Dulicecai()
        # 求往期和值尾
        a_hewei = (row[11].value + row[12].value + row[13].value) % 10  # 前3
        b_hewei = (row[8].value + row[9].value + row[10].value) % 10  # 前2
        c_hewei = (row[2].value + row[3].value + row[4].value) % 10  # 前1
        # 调用方法
        shu = a.hezhiwuxingshahaofa(a_hewei, b_hewei, c_hewei)
        # 选取符合预测和值尾的奖号
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            hezhi = item[0] + item[1] + item[2]
            hezhiwei = hezhi // 100 + hezhi // 10 + hezhi % 10
            if hezhiwei in shu:  # 如果符合和值尾
                if item not in All_kebian:
                    All_kebian.append(item)  # 选定

    def liuyao_danqishikongfa(self, row, All_kebian):
        dt = row[0].value + datetime.timedelta(hours=22)
        All_duizhao = All_kebian.copy()
        # 开奖号换算成动爻：百位是哪个数字即几爻动，十位在百位动爻的基础上向上行走，个位在十位动爻的基础上向上行走
        # print(row[2].value)
        # print(row[3].value)
        # print(row[4].value)
        if row[2].value == 0:
            yao1 = 4
        else:
            yao1 = row[2].value % 6
        if row[3].value == 0:
            yao2 = (yao1 + row[3].value + 10) % 6
        else:
            yao2 = (yao1 + row[3].value) % 6
        if row[4].value == 0:
            yao3 = (yao2 + row[4].value + 10) % 6
        else:
            yao3 = (yao2 + row[4].value) % 6
        if yao1 == 0:
            yao1 += 6
        if yao2 == 0:
            yao2 += 6
        if yao3 == 0:
            yao3 += 6
        a = liuyao.liuyao_api.Api()
        shu = []
        # 输入动爻起时间卦，返回动爻地支数
        # print(yao1)
        # print(yao2)
        # print(yao3)
        shu1 = a.get_danqishikongfa(dt, yao1)
        shu2 = a.get_danqishikongfa(dt, yao2)
        shu3 = a.get_danqishikongfa(dt, yao3)
        for i in shu1:
            if i not in shu:
                shu.append(i)
        for i in shu2:
            if i not in shu:
                shu.append(i)
        for i in shu3:
            if i not in shu:
                shu.append(i)
        print(shu)
        # 选取符合预测数的奖号
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            if item[0] in shu and item[1] in shu and item[2] in shu:  # 定数胆
                if item not in All_kebian:
                    All_kebian.append(item)  # 选定

    def liuyao_chunshijianguafa(self, row, All_kebian):
        dt = row[0].value + datetime.timedelta(hours=22)
        All_duizhao = All_kebian.copy()
        a = liuyao.liuyao_api.Api()
        shu = a.get_chunshijianguafa(dt, row[1].value%10)
        # 选取符合预测数的奖号
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            if item[0] in shu and item[1] in shu and item[2] in shu:  # 定数胆
                if item not in All_kebian:
                    All_kebian.append(item)  # 选定


# class Tongji:
#     # 组六：72%；组三28%（组三：27%；豹子：1%）
#     # 小小大/小大大/奇奇偶/奇偶偶：75%；全大/全小/全奇/全偶：25%
#     # 组选连号：
#     # 和值7-20：83.2%；和值0-6/21-27：16.8%
#     # 大小跨度：
#     def __init__(self):
#         pass
#
#     def tongji(self, row, zuliu):
#         zuliu_copy = zuliu.copy()  # zuliu_copy建立索引和查询，zuliu保存结果
#         pass


class Gailvfenxi:
    def __init__(self):
        self.Prk = fenxi.Pr(10, 3, True)[1]  # 考虑排三直选
        self.Pk = fenxi.P(10, 3, True)[1]  # 不考虑排三直选
        self.Crk = fenxi.Cr(10, 3, True)[1]  # 考虑排三组选
        self.Ck = fenxi.C(10, 3, True)[1]  # 不考虑排三组选
        # 建立开奖的可能性集合
        # 组合220
        self.Cbaozi = []
        self.Czusan = []
        self.Czuliu = []
        for zu in self.Crk:
            if zu[0] == zu[1] == zu[2]:
                self.Cbaozi.append([zu[0], zu[1], zu[2]])
            elif zu[0] == zu[1] or zu[0] == zu[2] or zu[1] == zu[2]:
                self.Czusan.append([zu[0], zu[1], zu[2]])
            else:
                self.Czuliu.append([zu[0], zu[1], zu[2]])
        # 排列1000
        self.Pbaozi = []
        self.Pzusan = []
        self.Pzuliu = []
        for zu in self.Prk:
            if zu[0] == zu[1] == zu[2]:
                self.Pbaozi.append([zu[0], zu[1], zu[2]])
            elif zu[0] == zu[1] or zu[0] == zu[2] or zu[1] == zu[2]:
                self.Pzusan.append([zu[0], zu[1], zu[2]])
            else:
                self.Pzuliu.append([zu[0], zu[1], zu[2]])
        # self.output()

    def output(self):
        print('组选：'+str(len(self.Crk)))
        # print(self.Crk)
        print('组选豹子：'+str(len(self.Cbaozi)))
        # print(self.Cbaozi)
        print('组选组三：'+str(len(self.Czusan)))
        # print(self.Czusan)
        print('组选组六：'+str(len(self.Czuliu)))
        # print(self.Czuliu)
        print('直选：' + str(len(self.Prk)))
        # print(self.Prk)
        print('直选豹子：'+str(len(self.Pbaozi)))
        # print(self.Pbaozi)
        print('直选组三：'+str(len(self.Pzusan)))
        # print(self.Pzusan)
        print('直选组六：'+str(len(self.Pzuliu)))
        # print(self.Pzuliu)
        print('豹子概率： 10/1000= 1%')
        print('组三概率：270/1000=27%')
        print('组六概率：720/1000=72%')

    # def tongji(self, path=None):
    #     # 读取数据表格
    #     realpath = os.path.join(os.getcwd(), path)
    #     wb_origin = openpyxl.load_workbook(realpath)
    #     ws_origin = wb_origin.active
    #     zongshu = 0  # 开奖总数
    #     mingzhong = 0  # 命中的开奖数
    #     for i, row in enumerate(ws_origin):
    #         # 升序排列，跳过缺少试机号的行
    #         if i <= 512:
    #             continue
    #         print('第' + str(row[0].value) + '期：')
    #         zuliu = self.Czuliu.copy()  # 要建立一份copy，否则会直接改变原列表的内容
    #         t = Tongji()
    #         t.tongji(row, zuliu)
    #         # 统计各种类型开奖号的概率
    #         zongshu += 1
    #         if sorted([row[1].value, row[2].value, row[3].value]) in zuliu:  # 组六
    #             mingzhong += 1
    #             pass
    #         else:
    #             pass
    #     print('命中：'+str(mingzhong)+' 总数：'+str(zongshu))
    #     print('概率：'+str(mingzhong/zongshu))

    def zhixuan_gailv(self, path=None):
        # 读取数据表格
        realpath = os.path.join(os.getcwd(), path)
        wb_origin = openpyxl.load_workbook(realpath)
        ws_origin = wb_origin.active
        zongshu = 0  # 开奖总数
        mingzhong = 0  # 命中的开奖数
        touzhu = 0  # 投注的累计注数
        jiangjin = 0  # 累计奖金数
        beitouflag = 1  # 倍投标记
        jieduantouru = 0  # 一直未中奖的阶段累计投入
        for i, row in enumerate(ws_origin):
            # 升序排列，跳过缺少试机号的行
            if i <= 512:
                continue
            # 测试
            if i > 5000:
                break
            print('第' + str(row[1].value) + '期：')
            Prk_kebian = self.Prk.copy()  # 要建立一份copy，否则会直接改变原列表的内容。后面直接对其操作，不需要返回值
            # # 进行定胆操作
            d = Xuanhaodingdanfangfa()
            # d.dingdan(row, Prk_kebian)
            # # 进行杀号操作
            s = Xuanhaoshahaofangfa()
            # s.shahao(row, Prk_kebian)
            # 进行数术操作
            ss = Shushufangfa()
            # ss.xiaochengtu_canwuyishufa(row, Prk_kebian)
            # ss.xiaochengtu_dingweidanfa(row, Prk_kebian)
            # ss.dulicecai_wuxingheyifa(row, Prk_kebian)
            # ss.dulicecai_hezhiwuxingshahaofa(row, Prk_kebian)
            # ss.liuyao_danqishikongfa(row, Prk_kebian)
            # 判断是否命中，并进行各类统计
            zongshu += 1  # 统计总数
            bencitouzhu = len(Prk_kebian)  # 本次投注注数
            touzhu += bencitouzhu * beitouflag  # 累计投注注数
            jieduantouru += bencitouzhu * beitouflag * 2  # 阶段投入钱数
            # 倍投算法：倍投后的奖金必须大于前面多次未中奖的投入包括本次投入
            while 173 * beitouflag < jieduantouru:
                beitouflag += 1
                jieduantouru += bencitouzhu * 2
                if beitouflag > 20:
                    print('陷入死循环，奖金永远小于投入，此法不能盈利，后面的计算结果无意义！')
                    break
            if (row[26].value, row[27].value, row[28].value) in Prk_kebian:  # 如果下期开奖号在投注集合里，则说明中奖
                mingzhong += 1
                jiangjin += 173 * beitouflag
                print('==========命中！=============')
                print('本次奖金' + str(173 * beitouflag))
                beitouflag = 1  # 如果中奖则阶段统计复位
                jieduantouru = 0  # 如果中奖则阶段统计复位
                print('阶段总结：投资：' + str(touzhu * 2) + ' 奖金：' + str(jiangjin))
            else:
                pass
            print(str(bencitouzhu) + '注。')
            if beitouflag > 1:
                print('倍投' + str(bencitouzhu * beitouflag) + '注。')
        print('命中：'+str(mingzhong)+' 总数：'+str(zongshu)+' 投资：'+str(touzhu*2)+' 奖金：'+str(jiangjin) + ' 投资回报率：'+str((jiangjin-touzhu*2)/(touzhu*2)))
        print('中奖概率：'+str(mingzhong/zongshu))

    def zhixuan_qujiangailv(self, qishiqi, qujianchangdu, kuadu):
        pass

    def zuxuan_gailv(self, path=None):
        # 读取数据表格
        realpath = os.path.join(os.getcwd(), path)
        wb_origin = openpyxl.load_workbook(realpath)
        ws_origin = wb_origin.active
        zongshu = 0  # 开奖总数
        mingzhong = 0  # 命中的开奖数
        touzhu = 0  # 投注的累计注数
        jiangjin = 0  # 累计奖金数
        beitouflag = 1  # 倍投标记
        jieduantouru = 0  # 一直未中奖的阶段累计投入
        for i, row in enumerate(ws_origin):
            # 升序排列，跳过缺少试机号的行
            if i <= 512:
                continue
            # 测试
            if i > 4000:
                break
            print('第' + str(row[1].value) + '期：')
            zuliu = self.Czuliu.copy()  # 要建立一份copy，否则会直接改变原列表的内容
            # # 进行定胆操作
            d = Xuanhaodingdanfangfa()
            # d.dingdan(row, zuliu)
            # # 进行杀号操作
            s = Xuanhaoshahaofangfa()
            # s.shahao(row, zuliu)
            # 进行数术操作
            ss = Shushufangfa()
            # ss.xiaochengtu_canwuyishufa(row, zuliu)
            # ss.xiaochengtu_dingweidanfa(row, zuliu)
            # ss.dulicecai_wuxingheyifa(row, zuliu)
            # ss.dulicecai_hezhiwuxingshahaofa(row, zuliu)
            # ss.liuyao_danqishikongfa(row, zuliu)
            # 判断是否命中，并进行各类统计
            zongshu += 1  # 统计总数
            bencitouzhu = len(zuliu)  # 本次投注注数
            touzhu += bencitouzhu * beitouflag  # 累计投注注数
            jieduantouru += bencitouzhu * beitouflag * 2  # 阶段投入钱数
            # 倍投算法：倍投后的奖金必须大于前面多次未中奖的投入包括本次投入
            while 173 * beitouflag < jieduantouru:
                beitouflag += 1
                jieduantouru += bencitouzhu * 2
                if beitouflag > 20:
                    print('陷入死循环，奖金永远小于投入，此法不能盈利，后面的计算结果无意义！')
                    break
            if sorted([row[26].value, row[27].value, row[28].value]) in zuliu:  # 排列组合已经自动从小到大排序，判断的时候也要排序
                mingzhong += 1
                jiangjin += 173 * beitouflag
                print('==========命中！=============')
                print('本次奖金'+str(173*beitouflag))
                beitouflag = 1  # 如果中奖则阶段统计复位
                jieduantouru = 0  # 如果中奖则阶段统计复位
                print('阶段总结：投资：'+str(touzhu*2)+' 奖金：'+str(jiangjin))
            else:
                pass
            print(str(bencitouzhu)+'注。')
            if beitouflag > 1:
                print('倍投' + str(bencitouzhu * beitouflag) + '注。')
        print('命中：'+str(mingzhong)+' 总数：'+str(zongshu)+' 投资：'+str(touzhu*2)+' 奖金：'+str(jiangjin) + ' 投资回报率：'+str((jiangjin-touzhu*2)/(touzhu*2)))
        print('中奖概率：'+str(mingzhong/zongshu))

    def zuxuan_qujiangailv(self, qishiqi, qujianchangdu, kuadu):
        pass
