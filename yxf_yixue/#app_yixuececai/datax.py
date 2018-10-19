# build-in
import os

# site-packages
import openpyxl
import openpyxl.utils


class Shuli:
    def __init__(self):
        pass

    def execute_xlsx1(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_1(realpath2)")  # 根据输入参数动态选取对应函数
        # 每次循环对应原始表格和目标表格的一行
        for i, row in enumerate(ws_origin):
            rows = ws_origin.max_row
            # 第一行为标题，需跳过
            if i == 0:
                continue
            # 循环到最末6行退出，防止有需要往期数据的情况
            if rows - i <= 6:
                break
            # for循环计数从0开始，但表格行列数从1开始，要注意差别
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_2(ws_origin, i)")  # 根据输入参数动态选取对应函数
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    def execute_xlsx2(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_3(ws_origin, realpath2)")
        for i, row in enumerate(ws_origin):
            # 表中有标题行，数据第1行无法取未来值，需跳过前2次循环
            if i <= 1:
                continue
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_4(ws_origin, i)")
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    @staticmethod
    def check(typ):
        if typ == "福彩3D":
            return "fucai3d"
        elif typ == "体彩排列三":
            return "ticaipailiesan"
        elif typ == "福彩双色球":
            return "fucaishuangseqiu"
        elif typ == "福彩七乐彩":
            return "fucaiqilecai"
        elif typ == "体彩大乐透":
            return "ticaidaletou"
        else:
            raise ImportError("输入参数错误：typ=" + typ)

    def fucai3d_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['开奖日期',
             '期号',
             '百位',
             '十位',
             '个位',
             '试机号百位',
             '试机号十位',
             '试机号个位',
             '前1期百位',
             '前1期十位',
             '前1期个位',
             '前2期百位',
             '前2期十位',
             '前2期个位',
             '前3期百位',
             '前3期十位',
             '前3期个位',
             '前4期百位',
             '前4期十位',
             '前4期个位',
             '前5期百位',
             '前5期十位',
             '前5期个位',
             '前6期百位',
             '前6期十位',
             '前6期个位'
             ])
        return wb_work, ws_work

    def fucai3d_2(self, ws_origin, i):
        # 当期数据
        arr = []
        KJRQ = ws_origin[i+1][0].value
        QH = ws_origin[i+1][1].value
        arr.append(KJRQ)
        arr.append(QH)
        BW = ws_origin[i+1][3].value
        SW = ws_origin[i+1][4].value
        GW = ws_origin[i+1][5].value
        sjh = str(ws_origin[i+1][6].value).zfill(3)
        if sjh == 'None':
            SJHBW = None
            SJHSW = None
            SJHGW = None
        else:
            SJHBW = int(sjh[0:1])
            SJHSW = int(sjh[1:2])
            SJHGW = int(sjh[2:3])
        Q1BW = ws_origin[i+2][3].value
        Q1SW = ws_origin[i+2][4].value
        Q1GW = ws_origin[i+2][5].value
        Q2BW = ws_origin[i+3][3].value
        Q2SW = ws_origin[i+3][4].value
        Q2GW = ws_origin[i+3][5].value
        Q3BW = ws_origin[i+4][3].value
        Q3SW = ws_origin[i+4][4].value
        Q3GW = ws_origin[i+4][5].value
        Q4BW = ws_origin[i+5][3].value
        Q4SW = ws_origin[i+5][4].value
        Q4GW = ws_origin[i+5][5].value
        Q5BW = ws_origin[i+6][3].value
        Q5SW = ws_origin[i+6][4].value
        Q5GW = ws_origin[i+6][5].value
        Q6BW = ws_origin[i+7][3].value
        Q6SW = ws_origin[i+7][4].value
        Q6GW = ws_origin[i+7][5].value
        arr.extend([BW, SW, GW, SJHBW, SJHSW, SJHGW, Q1BW, Q1SW, Q1GW, Q2BW, Q2SW, Q2GW, Q3BW, Q3SW, Q3GW, Q4BW, Q4SW, Q4GW, Q5BW, Q5SW, Q5GW, Q6BW, Q6SW, Q6GW])
        return arr

    def fucai3d_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.extend(['下期百位', '下期十位', '下期个位'])
        ws_work.append(arr)
        return wb_work, ws_work

    def fucai3d_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        XQBW = int(ws_origin[i][2].value)
        XQSW = int(ws_origin[i][3].value)
        XQGW = int(ws_origin[i][4].value)
        arr.extend([XQBW, XQSW, XQGW])
        return arr

    def ticaipailiesan_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['开奖日期',
             '期号',
             '百位',
             '十位',
             '个位',
             '前1期百位',
             '前1期十位',
             '前1期个位',
             '前2期百位',
             '前2期十位',
             '前2期个位',
             '前3期百位',
             '前3期十位',
             '前3期个位',
             '前4期百位',
             '前4期十位',
             '前4期个位',
             '前5期百位',
             '前5期十位',
             '前5期个位',
             '前6期百位',
             '前6期十位',
             '前6期个位'
             ])
        return wb_work, ws_work

    def ticaipailiesan_2(self, ws_origin, i):
        # 当期数据
        arr = []
        KJRQ = ws_origin[i+1][0].value
        QH = ws_origin[i+1][1].value
        arr.append(KJRQ)
        arr.append(QH)
        BW = ws_origin[i+1][3].value
        SW = ws_origin[i+1][4].value
        GW = ws_origin[i+1][5].value
        Q1BW = ws_origin[i+2][3].value
        Q1SW = ws_origin[i+2][4].value
        Q1GW = ws_origin[i+2][5].value
        Q2BW = ws_origin[i+3][3].value
        Q2SW = ws_origin[i+3][4].value
        Q2GW = ws_origin[i+3][5].value
        Q3BW = ws_origin[i+4][3].value
        Q3SW = ws_origin[i+4][4].value
        Q3GW = ws_origin[i+4][5].value
        Q4BW = ws_origin[i+5][3].value
        Q4SW = ws_origin[i+5][4].value
        Q4GW = ws_origin[i+5][5].value
        Q5BW = ws_origin[i+6][3].value
        Q5SW = ws_origin[i+6][4].value
        Q5GW = ws_origin[i+6][5].value
        Q6BW = ws_origin[i+7][3].value
        Q6SW = ws_origin[i+7][4].value
        Q6GW = ws_origin[i+7][5].value
        arr.extend([BW, SW, GW, Q1BW, Q1SW, Q1GW, Q2BW, Q2SW, Q2GW, Q3BW, Q3SW, Q3GW, Q4BW, Q4SW, Q4GW, Q5BW, Q5SW, Q5GW, Q6BW, Q6SW, Q6GW])
        return arr

    def ticaipailiesan_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.extend(['下期百位', '下期十位', '下期个位'])
        ws_work.append(arr)
        return wb_work, ws_work

    def ticaipailiesan_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        XQBW = int(ws_origin[i][2].value)
        XQSW = int(ws_origin[i][3].value)
        XQGW = int(ws_origin[i][4].value)
        arr.extend([XQBW, XQSW, XQGW])
        return arr


class WangqiLuyao:
    def __init__(self):
        pass

    def execute_xlsx1(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_1(realpath2)")  # 根据输入参数动态选取对应函数
        # 每次循环对应原始表格和目标表格的一行
        for i, row in enumerate(ws_origin):
            rows = ws_origin.max_row
            # 第一行为标题，需跳过
            if i == 0:
                continue
            # 循环到最末6行退出，防止计算六爻时出现空值
            if rows - i <= 6:
                break
            # for循环计数从0开始，但表格行列数从1开始，要注意差别
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_2(ws_origin, i)")  # 根据输入参数动态选取对应函数
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    def execute_xlsx2(self, typ, path1, path2):
        realpath1 = os.path.join(os.getcwd(), path1)
        realpath2 = os.path.join(os.getcwd(), path2)
        # 读取原始表格
        wb_origin = openpyxl.load_workbook(realpath1)
        ws_origin = wb_origin.active
        # 初始化目标表
        wb_work, ws_work = eval("self." + self.check(typ) + "_3(ws_origin, realpath2)")
        for i, row in enumerate(ws_origin):
            # 表中有标题行，数据第1行无法取未来值，需跳过前2次循环
            if i <= 1:
                continue
            # 数据列处理
            arr = eval("self." + self.check(typ) + "_4(ws_origin, i)")
            # 输出到目标表
            print(arr[0])
            ws_work.append(arr)
        # 保存文件
        wb_work.save(realpath2)

    @staticmethod
    def check(typ):
        if typ == "福彩3D":
            return "fucai3d"
        elif typ == "体彩排列三":
            return "ticaipailiesan"
        elif typ == "福彩双色球":
            return "fucaishuangseqiu"
        elif typ == "福彩七乐彩":
            return "fucaiqilecai"
        elif typ == "体彩大乐透":
            return "ticaidaletou"
        else:
            raise ImportError("输入参数错误：typ=" + typ)

    def fucai3d_1(self, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        ws_work.append(
            ['QH',  # 1,A
             'GH1',  # 2,B
             'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
             'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
             'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
             'F1',  # 29,AC
             'GH2',  # 30,AD
             'X21', 'X22', 'X23', 'X24', 'X25', 'X26',  # 31-36,AE-AJ
             'X27', 'X28', 'X29', 'X210', 'X211', 'X212', 'X213', 'X214', 'X215', 'X216',  # 37-46,AK-AT
             'X217', 'X218', 'X219', 'X220', 'X221', 'X222', 'X223', 'X224', 'X225', 'X226',  # 47-56,AU-BD
             'F2',  # 57,BE
             'GH3',  # 58,BF
             'X31', 'X32', 'X33', 'X34', 'X35', 'X36',  # 59-64,BG-BL
             'X37', 'X38', 'X39', 'X310', 'X311', 'X312', 'X313', 'X314', 'X315', 'X316',  # 65-74,BM-BV
             'X317', 'X318', 'X319', 'X320', 'X321', 'X322', 'X323', 'X324', 'X325', 'X326',  # 75-84,BW-CF
             'F3'  # 85,CG
             ])
        return wb_work, ws_work

    def fucai3d_2(self, ws_origin, i):
        # 当期数据
        arr = []
        QH = ws_origin[i+1][1].value
        arr.append(QH)
        for wei in range(3):
            locals()['F' + str(wei + 1)] = ws_origin[i + 1][3 + wei].value
            locals()['X' + str(wei + 1) + '1'] = ws_origin[i + 1][3 + wei].value - ws_origin[i + 2][3 + wei].value
            locals()['X' + str(wei + 1) + '2'] = ws_origin[i + 2][3 + wei].value - ws_origin[i + 3][3 + wei].value
            locals()['X' + str(wei + 1) + '3'] = ws_origin[i + 3][3 + wei].value - ws_origin[i + 4][3 + wei].value
            locals()['X' + str(wei + 1) + '4'] = ws_origin[i + 4][3 + wei].value - ws_origin[i + 5][3 + wei].value
            locals()['X' + str(wei + 1) + '5'] = ws_origin[i + 5][3 + wei].value - ws_origin[i + 6][3 + wei].value
            locals()['X' + str(wei + 1) + '6'] = ws_origin[i + 6][3 + wei].value - ws_origin[i + 7][3 + wei].value
            locals()['X' + str(wei + 1) + '7'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2']
            locals()['X' + str(wei + 1) + '8'] = locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
            locals()['X' + str(wei + 1) + '9'] = locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
            locals()['X' + str(wei + 1) + '10'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3']
            locals()['X' + str(wei + 1) + '11'] = locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
            locals()['X' + str(wei + 1) + '12'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
            locals()['X' + str(wei + 1) + '13'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
            locals()['X' + str(wei + 1) + '14'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
            locals()['X' + str(wei + 1) + '15'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
            locals()['X' + str(wei + 1) + '16'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
            locals()['X' + str(wei + 1) + '17'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
            locals()['X' + str(wei + 1) + '18'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
            locals()['X' + str(wei + 1) + '19'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
            locals()['X' + str(wei + 1) + '20'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
            locals()['X' + str(wei + 1) + '21'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
            locals()['X' + str(wei + 1) + '22'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
            locals()['X' + str(wei + 1) + '23'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
            locals()['X' + str(wei + 1) + '24'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
            locals()['X' + str(wei + 1) + '25'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5']
            locals()['X' + str(wei + 1) + '26'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
            # 六爻数值变换
            # 0和正变1，负变0
            locals()['X' + str(wei + 1) + '1g'] = 1 if locals()['X' + str(wei + 1) + '1'] >= 0 else 0
            locals()['X' + str(wei + 1) + '2g'] = 1 if locals()['X' + str(wei + 1) + '2'] >= 0 else 0
            locals()['X' + str(wei + 1) + '3g'] = 1 if locals()['X' + str(wei + 1) + '3'] >= 0 else 0
            locals()['X' + str(wei + 1) + '4g'] = 1 if locals()['X' + str(wei + 1) + '4'] >= 0 else 0
            locals()['X' + str(wei + 1) + '5g'] = 1 if locals()['X' + str(wei + 1) + '5'] >= 0 else 0
            locals()['X' + str(wei + 1) + '6g'] = 1 if locals()['X' + str(wei + 1) + '6'] >= 0 else 0
            # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
            locals()['guashu' + str(wei + 1)] = locals()['X' + str(wei + 1) + '1g']*1 + locals()['X' + str(wei + 1) + '2g']*2 + locals()['X' + str(wei + 1) + '3g']*4 + locals()['X' + str(wei + 1) + '4g']*8 + locals()['X' + str(wei + 1) + '5g']*16 + locals()['X' + str(wei + 1) + '6g']*32
            # 读取卦表
            wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\manual_data\基础表-六十四卦.xlsx'))
            ws_gua = wb_gua.active
            locals()['GH' + str(wei + 1)] = None
            # 查询卦表，求取卦号.注意卦数0-63，仅仅用于计算得到六十四卦，后面的格式化表等均用卦号1-64
            for row in ws_gua:
                if row[4].value == locals()['guashu' + str(wei + 1)]:
                    locals()['GH' + str(wei + 1)] = row[0].value
            arr.extend([
                locals()['GH' + str(wei + 1)],
                locals()['X' + str(wei + 1) + '1'],
                locals()['X' + str(wei + 1) + '2'],
                locals()['X' + str(wei + 1) + '3'],
                locals()['X' + str(wei + 1) + '4'],
                locals()['X' + str(wei + 1) + '5'],
                locals()['X' + str(wei + 1) + '6'],
                locals()['X' + str(wei + 1) + '7'],
                locals()['X' + str(wei + 1) + '8'],
                locals()['X' + str(wei + 1) + '9'],
                locals()['X' + str(wei + 1) + '10'],
                locals()['X' + str(wei + 1) + '11'],
                locals()['X' + str(wei + 1) + '12'],
                locals()['X' + str(wei + 1) + '13'],
                locals()['X' + str(wei + 1) + '14'],
                locals()['X' + str(wei + 1) + '15'],
                locals()['X' + str(wei + 1) + '16'],
                locals()['X' + str(wei + 1) + '17'],
                locals()['X' + str(wei + 1) + '18'],
                locals()['X' + str(wei + 1) + '19'],
                locals()['X' + str(wei + 1) + '20'],
                locals()['X' + str(wei + 1) + '21'],
                locals()['X' + str(wei + 1) + '22'],
                locals()['X' + str(wei + 1) + '23'],
                locals()['X' + str(wei + 1) + '24'],
                locals()['X' + str(wei + 1) + '25'],
                locals()['X' + str(wei + 1) + '26'],
                locals()['F' + str(wei + 1)]])
        return arr

    def fucai3d_3(self, ws_origin, realpath2):
        # 若已存在目标表，删除
        if os.path.exists(realpath2):
            os.remove(realpath2)
        # 重新建立目标表格
        wb_work = openpyxl.Workbook()
        ws_work = wb_work.active
        # 输出标题行
        arr = []
        for col in ws_origin[1]:
            arr.append(col.value)
        arr.extend(['R1', 'XFR1', 'R2', 'XFR2', 'R3', 'XFR3'])
        ws_work.append(arr)
        return wb_work, ws_work

    def fucai3d_4(self, ws_origin, i):
        # 因i从0开始，所以i+1为当期，i为未来期
        arr = []
        for col in ws_origin[i+1]:
            arr.append(col.value)
        for wei in range(3):
            locals()['R' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)
            locals()['XFR' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)-int(ws_origin[i+1][28 + wei*28].value)
            arr.extend([
                locals()['R' + str(wei + 1)],
                locals()['XFR' + str(wei + 1)]])
        return arr

    # def fucaiqilecai_1(self, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     ws_work.append(
    #         ['QH',  # 1,A
    #          'GH1',  # 2,B
    #          'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
    #          'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
    #          'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
    #          'F1',  # 29,AC
    #          'GH2',  # 30,AD
    #          'X21', 'X22', 'X23', 'X24', 'X25', 'X26',  # 31-36,AE-AJ
    #          'X27', 'X28', 'X29', 'X210', 'X211', 'X212', 'X213', 'X214', 'X215', 'X216',  # 37-46,AK-AT
    #          'X217', 'X218', 'X219', 'X220', 'X221', 'X222', 'X223', 'X224', 'X225', 'X226',  # 47-56,AU-BD
    #          'F2',  # 57,BE
    #          'GH3',  # 58,BF
    #          'X31', 'X32', 'X33', 'X34', 'X35', 'X36',  # 59-64,BG-BL
    #          'X37', 'X38', 'X39', 'X310', 'X311', 'X312', 'X313', 'X314', 'X315', 'X316',  # 65-74,BM-BV
    #          'X317', 'X318', 'X319', 'X320', 'X321', 'X322', 'X323', 'X324', 'X325', 'X326',  # 75-84,BW-CF
    #          'F3',  # 85,CG
    #          'GH4',  # 86
    #          'X41', 'X42', 'X43', 'X44', 'X45', 'X46',  # 87-92
    #          'X47', 'X48', 'X49', 'X410', 'X411', 'X412', 'X413', 'X414', 'X415', 'X416',  # 93-102
    #          'X417', 'X418', 'X419', 'X420', 'X421', 'X422', 'X423', 'X424', 'X425', 'X426',  # 103-112
    #          'F4',  # 113
    #          'GH5',  # 114
    #          'X51', 'X52', 'X53', 'X54', 'X55', 'X56',  # 115-120
    #          'X57', 'X58', 'X59', 'X510', 'X511', 'X512', 'X513', 'X514', 'X515', 'X516',  # 121-130
    #          'X517', 'X518', 'X519', 'X520', 'X521', 'X522', 'X523', 'X524', 'X525', 'X526',  # 131-140
    #          'F5',  # 141
    #          'GH6',  # 142
    #          'X61', 'X62', 'X63', 'X64', 'X65', 'X66',  # 143-148
    #          'X67', 'X68', 'X69', 'X610', 'X611', 'X612', 'X613', 'X614', 'X615', 'X616',  # 149-158
    #          'X617', 'X618', 'X619', 'X620', 'X621', 'X622', 'X623', 'X624', 'X625', 'X626',  # 159-168
    #          'F6',  # 169
    #          'GH7',  # 170
    #          'X71', 'X72', 'X73', 'X74', 'X75', 'X76',  # 171-176
    #          'X77', 'X78', 'X79', 'X710', 'X711', 'X712', 'X713', 'X714', 'X715', 'X716',  # 177-186
    #          'X717', 'X718', 'X719', 'X720', 'X721', 'X722', 'X723', 'X724', 'X725', 'X726',  # 187-196
    #          'F7',  # 197
    #          'GH8',  # 198
    #          'X81', 'X82', 'X83', 'X84', 'X85', 'X86',  # 199-204
    #          'X87', 'X88', 'X89', 'X810', 'X811', 'X812', 'X813', 'X814', 'X815', 'X816',  # 205-214
    #          'X817', 'X818', 'X819', 'X820', 'X821', 'X822', 'X823', 'X824', 'X825', 'X826',  # 215-224
    #          'F8'  # 225
    #          ])
    #     return wb_work, ws_work
    #
    # def fucaiqilecai_2(self, ws_origin, i):
    #     # 当期数据
    #     arr = []
    #     QH = ws_origin[i+1][1].value
    #     arr.append(QH)
    #     for wei in range(8):
    #         locals()['F' + str(wei + 1)] = ws_origin[i + 1][3 + wei].value
    #         locals()['X' + str(wei + 1) + '1'] = ws_origin[i + 1][3 + wei].value - ws_origin[i + 2][3 + wei].value
    #         locals()['X' + str(wei + 1) + '2'] = ws_origin[i + 2][3 + wei].value - ws_origin[i + 3][3 + wei].value
    #         locals()['X' + str(wei + 1) + '3'] = ws_origin[i + 3][3 + wei].value - ws_origin[i + 4][3 + wei].value
    #         locals()['X' + str(wei + 1) + '4'] = ws_origin[i + 4][3 + wei].value - ws_origin[i + 5][3 + wei].value
    #         locals()['X' + str(wei + 1) + '5'] = ws_origin[i + 5][3 + wei].value - ws_origin[i + 6][3 + wei].value
    #         locals()['X' + str(wei + 1) + '6'] = ws_origin[i + 6][3 + wei].value - ws_origin[i + 7][3 + wei].value
    #         locals()['X' + str(wei + 1) + '7'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '8'] = locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '9'] = locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '10'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '11'] = locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '12'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '13'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '14'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '15'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '16'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '17'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '18'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '19'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '20'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '21'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '22'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '23'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '24'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '25'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '26'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         # 六爻数值变换
    #         # 0和正变1，负变0
    #         locals()['X' + str(wei + 1) + '1g'] = 1 if locals()['X' + str(wei + 1) + '1'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '2g'] = 1 if locals()['X' + str(wei + 1) + '2'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '3g'] = 1 if locals()['X' + str(wei + 1) + '3'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '4g'] = 1 if locals()['X' + str(wei + 1) + '4'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '5g'] = 1 if locals()['X' + str(wei + 1) + '5'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '6g'] = 1 if locals()['X' + str(wei + 1) + '6'] >= 0 else 0
    #         # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
    #         locals()['guashu' + str(wei + 1)] = locals()['X' + str(wei + 1) + '1g']*1 + locals()['X' + str(wei + 1) + '2g']*2 + locals()['X' + str(wei + 1) + '3g']*4 + locals()['X' + str(wei + 1) + '4g']*8 + locals()['X' + str(wei + 1) + '5g']*16 + locals()['X' + str(wei + 1) + '6g']*32
    #         # 读取卦表
    #         wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\manual_data\基础表-六十四卦.xlsx'))
    #         ws_gua = wb_gua.active
    #         locals()['GH' + str(wei + 1)] = None
    #         # 查询卦表，求取卦号.注意卦数0-63，仅仅用于计算得到六十四卦，后面的格式化表等均用卦号1-64
    #         for row in ws_gua:
    #             if row[4].value == locals()['guashu' + str(wei + 1)]:
    #                 locals()['GH' + str(wei + 1)] = row[0].value
    #         arr.extend([
    #             locals()['GH' + str(wei + 1)],
    #             locals()['X' + str(wei + 1) + '1'],
    #             locals()['X' + str(wei + 1) + '2'],
    #             locals()['X' + str(wei + 1) + '3'],
    #             locals()['X' + str(wei + 1) + '4'],
    #             locals()['X' + str(wei + 1) + '5'],
    #             locals()['X' + str(wei + 1) + '6'],
    #             locals()['X' + str(wei + 1) + '7'],
    #             locals()['X' + str(wei + 1) + '8'],
    #             locals()['X' + str(wei + 1) + '9'],
    #             locals()['X' + str(wei + 1) + '10'],
    #             locals()['X' + str(wei + 1) + '11'],
    #             locals()['X' + str(wei + 1) + '12'],
    #             locals()['X' + str(wei + 1) + '13'],
    #             locals()['X' + str(wei + 1) + '14'],
    #             locals()['X' + str(wei + 1) + '15'],
    #             locals()['X' + str(wei + 1) + '16'],
    #             locals()['X' + str(wei + 1) + '17'],
    #             locals()['X' + str(wei + 1) + '18'],
    #             locals()['X' + str(wei + 1) + '19'],
    #             locals()['X' + str(wei + 1) + '20'],
    #             locals()['X' + str(wei + 1) + '21'],
    #             locals()['X' + str(wei + 1) + '22'],
    #             locals()['X' + str(wei + 1) + '23'],
    #             locals()['X' + str(wei + 1) + '24'],
    #             locals()['X' + str(wei + 1) + '25'],
    #             locals()['X' + str(wei + 1) + '26'],
    #             locals()['F' + str(wei + 1)]])
    #     return arr
    #
    # def fucaiqilecai_3(self, ws_origin, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     arr = []
    #     for col in ws_origin[1]:
    #         arr.append(col.value)
    #     arr.extend(['R1', 'XFR1', 'R2', 'XFR2', 'R3', 'XFR3', 'R4', 'XFR4', 'R5', 'XFR5', 'R6', 'XFR6', 'R7', 'XFR7', 'R8', 'XFR8'])
    #     ws_work.append(arr)
    #     return wb_work, ws_work
    #
    # def fucaiqilecai_4(self, ws_origin, i):
    #     # 因i从0开始，所以i+1为当期，i为未来期
    #     arr = []
    #     for col in ws_origin[i+1]:
    #         arr.append(col.value)
    #     for wei in range(8):
    #         locals()['R' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)
    #         locals()['XFR' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)-int(ws_origin[i+1][28 + wei*28].value)
    #         arr.extend([
    #             locals()['R' + str(wei + 1)],
    #             locals()['XFR' + str(wei + 1)]])
    #     return arr
    #
    # def fucaishuangseqiu_1(self, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     ws_work.append(
    #         ['QH',  # 1,A
    #          'GH1',  # 2,B
    #          'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
    #          'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
    #          'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
    #          'F1',  # 29,AC
    #          'GH2',  # 30,AD
    #          'X21', 'X22', 'X23', 'X24', 'X25', 'X26',  # 31-36,AE-AJ
    #          'X27', 'X28', 'X29', 'X210', 'X211', 'X212', 'X213', 'X214', 'X215', 'X216',  # 37-46,AK-AT
    #          'X217', 'X218', 'X219', 'X220', 'X221', 'X222', 'X223', 'X224', 'X225', 'X226',  # 47-56,AU-BD
    #          'F2',  # 57,BE
    #          'GH3',  # 58,BF
    #          'X31', 'X32', 'X33', 'X34', 'X35', 'X36',  # 59-64,BG-BL
    #          'X37', 'X38', 'X39', 'X310', 'X311', 'X312', 'X313', 'X314', 'X315', 'X316',  # 65-74,BM-BV
    #          'X317', 'X318', 'X319', 'X320', 'X321', 'X322', 'X323', 'X324', 'X325', 'X326',  # 75-84,BW-CF
    #          'F3',  # 85,CG
    #          'GH4',  # 86
    #          'X41', 'X42', 'X43', 'X44', 'X45', 'X46',  # 87-92
    #          'X47', 'X48', 'X49', 'X410', 'X411', 'X412', 'X413', 'X414', 'X415', 'X416',  # 93-102
    #          'X417', 'X418', 'X419', 'X420', 'X421', 'X422', 'X423', 'X424', 'X425', 'X426',  # 103-112
    #          'F4',  # 113
    #          'GH5',  # 114
    #          'X51', 'X52', 'X53', 'X54', 'X55', 'X56',  # 115-120
    #          'X57', 'X58', 'X59', 'X510', 'X511', 'X512', 'X513', 'X514', 'X515', 'X516',  # 121-130
    #          'X517', 'X518', 'X519', 'X520', 'X521', 'X522', 'X523', 'X524', 'X525', 'X526',  # 131-140
    #          'F5',  # 141
    #          'GH6',  # 142
    #          'X61', 'X62', 'X63', 'X64', 'X65', 'X66',  # 143-148
    #          'X67', 'X68', 'X69', 'X610', 'X611', 'X612', 'X613', 'X614', 'X615', 'X616',  # 149-158
    #          'X617', 'X618', 'X619', 'X620', 'X621', 'X622', 'X623', 'X624', 'X625', 'X626',  # 159-168
    #          'F6',  # 169
    #          'GH7',  # 170
    #          'X71', 'X72', 'X73', 'X74', 'X75', 'X76',  # 171-176
    #          'X77', 'X78', 'X79', 'X710', 'X711', 'X712', 'X713', 'X714', 'X715', 'X716',  # 177-186
    #          'X717', 'X718', 'X719', 'X720', 'X721', 'X722', 'X723', 'X724', 'X725', 'X726',  # 187-196
    #          'F7'  # 197
    #          ])
    #     return wb_work, ws_work
    #
    # def fucaishuangseqiu_2(self, ws_origin, i):
    #     # 当期数据
    #     arr = []
    #     QH = ws_origin[i+1][1].value
    #     arr.append(QH)
    #     for wei in range(7):
    #         locals()['F' + str(wei + 1)] = ws_origin[i + 1][3 + wei].value
    #         locals()['X' + str(wei + 1) + '1'] = ws_origin[i + 1][2 + wei].value - ws_origin[i + 2][2 + wei].value
    #         locals()['X' + str(wei + 1) + '2'] = ws_origin[i + 2][2 + wei].value - ws_origin[i + 3][2 + wei].value
    #         locals()['X' + str(wei + 1) + '3'] = ws_origin[i + 3][2 + wei].value - ws_origin[i + 4][2 + wei].value
    #         locals()['X' + str(wei + 1) + '4'] = ws_origin[i + 4][2 + wei].value - ws_origin[i + 5][2 + wei].value
    #         locals()['X' + str(wei + 1) + '5'] = ws_origin[i + 5][2 + wei].value - ws_origin[i + 6][2 + wei].value
    #         locals()['X' + str(wei + 1) + '6'] = ws_origin[i + 6][2 + wei].value - ws_origin[i + 7][2 + wei].value
    #         locals()['X' + str(wei + 1) + '7'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '8'] = locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '9'] = locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '10'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '11'] = locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '12'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '13'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '14'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '15'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '16'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '17'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '18'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '19'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '20'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '21'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '22'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '23'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '24'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '25'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '26'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         # 六爻数值变换
    #         # 0和正变1，负变0
    #         locals()['X' + str(wei + 1) + '1g'] = 1 if locals()['X' + str(wei + 1) + '1'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '2g'] = 1 if locals()['X' + str(wei + 1) + '2'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '3g'] = 1 if locals()['X' + str(wei + 1) + '3'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '4g'] = 1 if locals()['X' + str(wei + 1) + '4'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '5g'] = 1 if locals()['X' + str(wei + 1) + '5'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '6g'] = 1 if locals()['X' + str(wei + 1) + '6'] >= 0 else 0
    #         # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
    #         locals()['guashu' + str(wei + 1)] = locals()['X' + str(wei + 1) + '1g']*1 + locals()['X' + str(wei + 1) + '2g']*2 + locals()['X' + str(wei + 1) + '3g']*4 + locals()['X' + str(wei + 1) + '4g']*8 + locals()['X' + str(wei + 1) + '5g']*16 + locals()['X' + str(wei + 1) + '6g']*32
    #         # 读取卦表
    #         wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\manual_data\基础表-六十四卦.xlsx'))
    #         ws_gua = wb_gua.active
    #         locals()['GH' + str(wei + 1)] = None
    #         # 查询卦表，求取卦号.注意卦数0-63，仅仅用于计算得到六十四卦，后面的格式化表等均用卦号1-64
    #         for row in ws_gua:
    #             if row[4].value == locals()['guashu' + str(wei + 1)]:
    #                 locals()['GH' + str(wei + 1)] = row[0].value
    #         arr.extend([
    #             locals()['GH' + str(wei + 1)],
    #             locals()['X' + str(wei + 1) + '1'],
    #             locals()['X' + str(wei + 1) + '2'],
    #             locals()['X' + str(wei + 1) + '3'],
    #             locals()['X' + str(wei + 1) + '4'],
    #             locals()['X' + str(wei + 1) + '5'],
    #             locals()['X' + str(wei + 1) + '6'],
    #             locals()['X' + str(wei + 1) + '7'],
    #             locals()['X' + str(wei + 1) + '8'],
    #             locals()['X' + str(wei + 1) + '9'],
    #             locals()['X' + str(wei + 1) + '10'],
    #             locals()['X' + str(wei + 1) + '11'],
    #             locals()['X' + str(wei + 1) + '12'],
    #             locals()['X' + str(wei + 1) + '13'],
    #             locals()['X' + str(wei + 1) + '14'],
    #             locals()['X' + str(wei + 1) + '15'],
    #             locals()['X' + str(wei + 1) + '16'],
    #             locals()['X' + str(wei + 1) + '17'],
    #             locals()['X' + str(wei + 1) + '18'],
    #             locals()['X' + str(wei + 1) + '19'],
    #             locals()['X' + str(wei + 1) + '20'],
    #             locals()['X' + str(wei + 1) + '21'],
    #             locals()['X' + str(wei + 1) + '22'],
    #             locals()['X' + str(wei + 1) + '23'],
    #             locals()['X' + str(wei + 1) + '24'],
    #             locals()['X' + str(wei + 1) + '25'],
    #             locals()['X' + str(wei + 1) + '26'],
    #             locals()['F' + str(wei + 1)]])
    #     return arr
    #
    # def fucaishuangseqiu_3(self, ws_origin, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     arr = []
    #     for col in ws_origin[1]:
    #         arr.append(col.value)
    #     arr.extend(
    #         ['R1', 'XFR1', 'R2', 'XFR2', 'R3', 'XFR3', 'R4', 'XFR4', 'R5', 'XFR5', 'R6', 'XFR6', 'R7', 'XFR7'])
    #     ws_work.append(arr)
    #     return wb_work, ws_work
    #
    # def fucaishuangseqiu_4(self, ws_origin, i):
    #     # 因i从0开始，所以i+1为当期，i为未来期
    #     arr = []
    #     for col in ws_origin[i+1]:
    #         arr.append(col.value)
    #     for wei in range(7):
    #         locals()['R' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)
    #         locals()['XFR' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)-int(ws_origin[i+1][28 + wei*28].value)
    #         arr.extend([
    #             locals()['R' + str(wei + 1)],
    #             locals()['XFR' + str(wei + 1)]])
    #     return arr
    #
    # def ticaidaletou_1(self, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     ws_work.append(
    #         ['QH',  # 1,A
    #          'GH1',  # 2,B
    #          'X11', 'X12', 'X13', 'X14', 'X15', 'X16',  # 3-8,C-H
    #          'X17', 'X18', 'X19', 'X110', 'X111', 'X112', 'X113', 'X114', 'X115', 'X116',  # 9-18,I-R
    #          'X117', 'X118', 'X119', 'X120', 'X121', 'X122', 'X123', 'X124', 'X125', 'X126',  # 19-28,S-AB
    #          'F1',  # 29,AC
    #          'GH2',  # 30,AD
    #          'X21', 'X22', 'X23', 'X24', 'X25', 'X26',  # 31-36,AE-AJ
    #          'X27', 'X28', 'X29', 'X210', 'X211', 'X212', 'X213', 'X214', 'X215', 'X216',  # 37-46,AK-AT
    #          'X217', 'X218', 'X219', 'X220', 'X221', 'X222', 'X223', 'X224', 'X225', 'X226',  # 47-56,AU-BD
    #          'F2',  # 57,BE
    #          'GH3',  # 58,BF
    #          'X31', 'X32', 'X33', 'X34', 'X35', 'X36',  # 59-64,BG-BL
    #          'X37', 'X38', 'X39', 'X310', 'X311', 'X312', 'X313', 'X314', 'X315', 'X316',  # 65-74,BM-BV
    #          'X317', 'X318', 'X319', 'X320', 'X321', 'X322', 'X323', 'X324', 'X325', 'X326',  # 75-84,BW-CF
    #          'F3',  # 85,CG
    #          'GH4',  # 86
    #          'X41', 'X42', 'X43', 'X44', 'X45', 'X46',  # 87-92
    #          'X47', 'X48', 'X49', 'X410', 'X411', 'X412', 'X413', 'X414', 'X415', 'X416',  # 93-102
    #          'X417', 'X418', 'X419', 'X420', 'X421', 'X422', 'X423', 'X424', 'X425', 'X426',  # 103-112
    #          'F4',  # 113
    #          'GH5',  # 114
    #          'X51', 'X52', 'X53', 'X54', 'X55', 'X56',  # 115-120
    #          'X57', 'X58', 'X59', 'X510', 'X511', 'X512', 'X513', 'X514', 'X515', 'X516',  # 121-130
    #          'X517', 'X518', 'X519', 'X520', 'X521', 'X522', 'X523', 'X524', 'X525', 'X526',  # 131-140
    #          'F5',  # 141
    #          'GH6',  # 142
    #          'X61', 'X62', 'X63', 'X64', 'X65', 'X66',  # 143-148
    #          'X67', 'X68', 'X69', 'X610', 'X611', 'X612', 'X613', 'X614', 'X615', 'X616',  # 149-158
    #          'X617', 'X618', 'X619', 'X620', 'X621', 'X622', 'X623', 'X624', 'X625', 'X626',  # 159-168
    #          'F6',  # 169
    #          'GH7',  # 170
    #          'X71', 'X72', 'X73', 'X74', 'X75', 'X76',  # 171-176
    #          'X77', 'X78', 'X79', 'X710', 'X711', 'X712', 'X713', 'X714', 'X715', 'X716',  # 177-186
    #          'X717', 'X718', 'X719', 'X720', 'X721', 'X722', 'X723', 'X724', 'X725', 'X726',  # 187-196
    #          'F7'  # 197
    #          ])
    #     return wb_work, ws_work
    #
    # def ticaidaletou_2(self, ws_origin, i):
    #     # 当期数据
    #     arr = []
    #     QH = ws_origin[i+1][1].value
    #     arr.append(QH)
    #     for wei in range(7):
    #         locals()['F' + str(wei + 1)] = ws_origin[i + 1][3 + wei].value
    #         locals()['X' + str(wei + 1) + '1'] = ws_origin[i + 1][2 + wei].value - ws_origin[i + 2][2 + wei].value
    #         locals()['X' + str(wei + 1) + '2'] = ws_origin[i + 2][2 + wei].value - ws_origin[i + 3][2 + wei].value
    #         locals()['X' + str(wei + 1) + '3'] = ws_origin[i + 3][2 + wei].value - ws_origin[i + 4][2 + wei].value
    #         locals()['X' + str(wei + 1) + '4'] = ws_origin[i + 4][2 + wei].value - ws_origin[i + 5][2 + wei].value
    #         locals()['X' + str(wei + 1) + '5'] = ws_origin[i + 5][2 + wei].value - ws_origin[i + 6][2 + wei].value
    #         locals()['X' + str(wei + 1) + '6'] = ws_origin[i + 6][2 + wei].value - ws_origin[i + 7][2 + wei].value
    #         locals()['X' + str(wei + 1) + '7'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '8'] = locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '9'] = locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '10'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '11'] = locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '12'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '13'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '14'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '15'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '16'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '17'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '18'] = locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1'] * locals()['X' + str(wei + 1) + '1']
    #         locals()['X' + str(wei + 1) + '19'] = locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2'] * locals()['X' + str(wei + 1) + '2']
    #         locals()['X' + str(wei + 1) + '20'] = locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3'] * locals()['X' + str(wei + 1) + '3']
    #         locals()['X' + str(wei + 1) + '21'] = locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4'] * locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '22'] = locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5'] * locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '23'] = locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6'] * locals()['X' + str(wei + 1) + '6']
    #         locals()['X' + str(wei + 1) + '24'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4']
    #         locals()['X' + str(wei + 1) + '25'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5']
    #         locals()['X' + str(wei + 1) + '26'] = locals()['X' + str(wei + 1) + '1'] + locals()['X' + str(wei + 1) + '2'] + locals()['X' + str(wei + 1) + '3'] + locals()['X' + str(wei + 1) + '4'] + locals()['X' + str(wei + 1) + '5'] + locals()['X' + str(wei + 1) + '6']
    #         # 六爻数值变换
    #         # 0和正变1，负变0
    #         locals()['X' + str(wei + 1) + '1g'] = 1 if locals()['X' + str(wei + 1) + '1'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '2g'] = 1 if locals()['X' + str(wei + 1) + '2'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '3g'] = 1 if locals()['X' + str(wei + 1) + '3'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '4g'] = 1 if locals()['X' + str(wei + 1) + '4'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '5g'] = 1 if locals()['X' + str(wei + 1) + '5'] >= 0 else 0
    #         locals()['X' + str(wei + 1) + '6g'] = 1 if locals()['X' + str(wei + 1) + '6'] >= 0 else 0
    #         # 以X11为初爻、低位，X16为上爻、高位的方法计算卦数
    #         locals()['guashu' + str(wei + 1)] = locals()['X' + str(wei + 1) + '1g']*1 + locals()['X' + str(wei + 1) + '2g']*2 + locals()['X' + str(wei + 1) + '3g']*4 + locals()['X' + str(wei + 1) + '4g']*8 + locals()['X' + str(wei + 1) + '5g']*16 + locals()['X' + str(wei + 1) + '6g']*32
    #         # 读取卦表
    #         wb_gua = openpyxl.load_workbook(os.path.join(os.getcwd(), '..\common\manual_data\基础表-六十四卦.xlsx'))
    #         ws_gua = wb_gua.active
    #         locals()['GH' + str(wei + 1)] = None
    #         # 查询卦表，求取卦号.注意卦数0-63，仅仅用于计算得到六十四卦，后面的格式化表等均用卦号1-64
    #         for row in ws_gua:
    #             if row[4].value == locals()['guashu' + str(wei + 1)]:
    #                 locals()['GH' + str(wei + 1)] = row[0].value
    #         arr.extend([
    #             locals()['GH' + str(wei + 1)],
    #             locals()['X' + str(wei + 1) + '1'],
    #             locals()['X' + str(wei + 1) + '2'],
    #             locals()['X' + str(wei + 1) + '3'],
    #             locals()['X' + str(wei + 1) + '4'],
    #             locals()['X' + str(wei + 1) + '5'],
    #             locals()['X' + str(wei + 1) + '6'],
    #             locals()['X' + str(wei + 1) + '7'],
    #             locals()['X' + str(wei + 1) + '8'],
    #             locals()['X' + str(wei + 1) + '9'],
    #             locals()['X' + str(wei + 1) + '10'],
    #             locals()['X' + str(wei + 1) + '11'],
    #             locals()['X' + str(wei + 1) + '12'],
    #             locals()['X' + str(wei + 1) + '13'],
    #             locals()['X' + str(wei + 1) + '14'],
    #             locals()['X' + str(wei + 1) + '15'],
    #             locals()['X' + str(wei + 1) + '16'],
    #             locals()['X' + str(wei + 1) + '17'],
    #             locals()['X' + str(wei + 1) + '18'],
    #             locals()['X' + str(wei + 1) + '19'],
    #             locals()['X' + str(wei + 1) + '20'],
    #             locals()['X' + str(wei + 1) + '21'],
    #             locals()['X' + str(wei + 1) + '22'],
    #             locals()['X' + str(wei + 1) + '23'],
    #             locals()['X' + str(wei + 1) + '24'],
    #             locals()['X' + str(wei + 1) + '25'],
    #             locals()['X' + str(wei + 1) + '26'],
    #             locals()['F' + str(wei + 1)]])
    #     return arr
    #
    # def ticaidaletou_3(self, ws_origin, realpath2):
    #     # 若已存在目标表，删除
    #     if os.path.exists(realpath2):
    #         os.remove(realpath2)
    #     # 重新建立目标表格
    #     wb_work = openpyxl.Workbook()
    #     ws_work = wb_work.active
    #     # 输出标题行
    #     arr = []
    #     for col in ws_origin[1]:
    #         arr.append(col.value)
    #     arr.extend(
    #         ['R1', 'XFR1', 'R2', 'XFR2', 'R3', 'XFR3', 'R4', 'XFR4', 'R5', 'XFR5', 'R6', 'XFR6', 'R7', 'XFR7'])
    #     ws_work.append(arr)
    #     return wb_work, ws_work
    #
    # def ticaidaletou_4(self, ws_origin, i):
    #     # 因i从0开始，所以i+1为当期，i为未来期
    #     arr = []
    #     for col in ws_origin[i+1]:
    #         arr.append(col.value)
    #     for wei in range(7):
    #         locals()['R' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)
    #         locals()['XFR' + str(wei + 1)] = int(ws_origin[i][28 + wei*28].value)-int(ws_origin[i+1][28 + wei*28].value)
    #         arr.extend([
    #             locals()['R' + str(wei + 1)],
    #             locals()['XFR' + str(wei + 1)]])
    #     return arr
