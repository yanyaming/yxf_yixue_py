"""
    分析回归模型:

"""

# build-in
import os


# site-packages
import openpyxl
import openpyxl.utils
import yixuececai.fenxi as fenxi


class Huiguifenxi:
    def __init__(self):
        self.Prk = fenxi.Pr(10, 3, True)[1]  # 考虑排三直选
        self.Pk = fenxi.P(10, 3, True)[1]  # 不考虑排三直选
        self.Crk = fenxi.Cr(10, 3, True)[1]  # 考虑排三组选
        self.Ck = fenxi.C(10, 3, True)[1]  # 不考虑排三组选
        # 建立开奖的可能性集合
        # 组合220
        self.Cbaozi = []
        self.Czusan = []
        self.Czuliu = []
        for zu in self.Crk:
            if zu[0] == zu[1] == zu[2]:
                self.Cbaozi.append([zu[0], zu[1], zu[2]])
            elif zu[0] == zu[1] or zu[0] == zu[2] or zu[1] == zu[2]:
                self.Czusan.append([zu[0], zu[1], zu[2]])
            else:
                self.Czuliu.append([zu[0], zu[1], zu[2]])
        # 排列1000
        self.Pbaozi = []
        self.Pzusan = []
        self.Pzuliu = []
        for zu in self.Prk:
            if zu[0] == zu[1] == zu[2]:
                self.Pbaozi.append([zu[0], zu[1], zu[2]])
            elif zu[0] == zu[1] or zu[0] == zu[2] or zu[1] == zu[2]:
                self.Pzusan.append([zu[0], zu[1], zu[2]])
            else:
                self.Pzuliu.append([zu[0], zu[1], zu[2]])
        # self.output()

    def read_moxing(self, row, All_kebian):
        All_duizhao = All_kebian.copy()
        data = []  # 存储本期开奖数据
        for i, col in enumerate(row):
            if 1 < i <= 28:
                data.append(col.value)
        print(data)
        realpath = os.path.join(os.getcwd(), "..\yixuececai\manual_data\福彩3D_往期六爻_回归模型.xlsx")
        ws_origin = openpyxl.load_workbook(realpath).active
        model = []  # 读取特定位卦的模型系数
        for j in range(2, 30):
            model.append(ws_origin[4][j].value)
        model.append(ws_origin[4][91].value)
        print(model)
        num = model[28]
        for i in range(0, 27):
            # print(model[i]*data[i])
            num += model[i]*data[i]
        print(round(num))
        num2 = round(num) + 1
        num3 = round(num) - 1
        if round(num) + 1 >= 10:
            num2 -= 10
        if round(num) - 1 < 0:
            num3 += 10
        All_kebian.clear()
        for item in All_duizhao:  # 每一项
            if item[0] == round(num) or item[0] == num2 or item[0] == num3:  # 定位胆
                if item not in All_kebian:
                    All_kebian.append(item)  # 选定
        # print(All_kebian)

    def ceshi(self):
        # 读取数据表格
        realpath = os.path.join(os.getcwd(), 'work_data\福彩3D_往期六爻2.xlsx')
        wb_origin = openpyxl.load_workbook(realpath)
        ws_origin = wb_origin.active
        zongshu = 0  # 开奖总数
        mingzhong = 0  # 命中的开奖数
        for i, row in enumerate(ws_origin):
            # 测试
            # if i < 1500:
            #     continue
            if i > 5000:
                break
            # 选取特定卦号
            if row[1].value == 3:
                print('第' + str(row[0].value) + '期：')
                Prk_kebian = self.Prk.copy()  # 要建立一份copy，否则会直接改变原列表的内容。后面直接对其操作，不需要返回值
                self.read_moxing(row, Prk_kebian)
                zongshu += 1  # 统计总数
                if (row[85].value, row[87].value, row[89].value) in Prk_kebian:  # 如果下期开奖号在投注集合里，则说明中奖
                    mingzhong += 1
                    print('==========命中！=============')
                else:
                    pass
        print('命中：' + str(mingzhong) + ' 总数：' + str(zongshu))
        print('概率：' + str(mingzhong / zongshu))


def fucai3d_1(realpath3):
    # 若已存在目标表，删除
    if os.path.exists(realpath3):
        os.remove(realpath3)
    # 重新建立目标表格
    wb_gailvfenxi = openpyxl.Workbook()
    ws_gailvfenxi = wb_gailvfenxi.active
    # 输出标题行
    ws_gailvfenxi.append(['NUM', 'GH1', 'GH2', 'GH3', 'PC1', 'PC2', 'PC3'])
    return wb_gailvfenxi, ws_gailvfenxi


def fucai3d_2(ws_geshihua, ws_huiguimoxing, i):
    # 取出开奖数据的卦号
    GH1 = ws_geshihua[i + 2][1].value
    GH2 = ws_geshihua[i + 2][29].value
    GH3 = ws_geshihua[i + 2][57].value
    # 计算速度太慢，可只单独分析一卦
    if GH1 == 4:
    #if True:
        # 对每一期进行分析，用期号索引
        NUM = ws_geshihua[i + 2][0].value
        # 下期升降的实际值
        XFR1 = ws_geshihua[i + 2][88].value
        XFR2 = ws_geshihua[i + 2][89].value
        XFR3 = ws_geshihua[i + 2][90].value
        # 下期升降的估计值。先加上模型中的常量
        XFR1p = ws_huiguimoxing[GH1 + 1][91].value
        XFR2p = ws_huiguimoxing[GH2 + 1][92].value
        XFR3p = ws_huiguimoxing[GH3 + 1][93].value
        # 后面各项需要两个表按照卦号匹配然后交叉相乘
        for j, col in enumerate(ws_geshihua[i + 2]):
            if 2 <= j <= 28:  # 百位
                XFR1p += (ws_geshihua[i + 2][j].value * ws_huiguimoxing[GH1 + 1][j].value)  # 匹配对应卦的回归系数
            if 30 <= j <= 56:  # 十位
                XFR2p += (ws_geshihua[i + 2][j].value * ws_huiguimoxing[GH2 + 1][j].value)  # 匹配对应卦的回归系数
            if 58 <= j <= 84:  # 个位
                XFR3p += (ws_geshihua[i + 2][j].value * ws_huiguimoxing[GH3 + 1][j].value)  # 匹配对应卦的回归系数
        # 先取整，再求出每一个开奖数据的偏差绝对值
        PC1 = abs(XFR1 - round(XFR1p))
        PC2 = abs(XFR2 - round(XFR2p))
        PC3 = abs(XFR3 - round(XFR3p))
        arr = [NUM, GH1, GH2, GH3, PC1, PC2, PC3]
    else:
        arr = 0
    return arr


def fucai3d_3(realpath3):
    wb_gailvfenxi = openpyxl.load_workbook(realpath3)
    ws_gailvfenxi = wb_gailvfenxi.active
    arr1 = []
    arr2 = []
    arr3 = []
    for i in range(64):
        # 每一个特定的卦和位，都要遍历一次卦-偏差对应表
        tongji_1 = 0
        tongji_10 = 0
        tongji_11 = 0
        tongji_12 = 0
        tongji_2 = 0
        tongji_20 = 0
        tongji_21 = 0
        tongji_22 = 0
        tongji_3 = 0
        tongji_30 = 0
        tongji_31 = 0
        tongji_32 = 0
        for j, row in enumerate(ws_gailvfenxi):
            if j <= 1:
                continue
            # 匹配
            if ws_gailvfenxi[j][1].value == i+1:
                # 百位偏差
                tongji_1 += 1  # 统计出现此卦的总数
                if ws_gailvfenxi[j][4].value == 0:  # 统计命中的个数
                    tongji_10 += 1
                    tongji_11 += 1
                    tongji_12 += 1
                if ws_gailvfenxi[j][4].value == 1:  # 统计命中的个数
                    tongji_11 += 1
                    tongji_12 += 1
                if ws_gailvfenxi[j][4].value == 2:  # 统计命中的个数
                    tongji_12 += 1
            if ws_gailvfenxi[j][2].value == i+1:
                # 十位偏差
                tongji_2 += 1  # 统计出现此卦的总数
                if ws_gailvfenxi[j][5].value == 0:  # 统计命中的个数
                    tongji_20 += 1
                    tongji_21 += 1
                    tongji_22 += 1
                if ws_gailvfenxi[j][5].value == 1:  # 统计命中的个数
                    tongji_21 += 1
                    tongji_22 += 1
                if ws_gailvfenxi[j][5].value == 2:  # 统计命中的个数
                    tongji_22 += 1
            if ws_gailvfenxi[j][3].value == i+1:
                # 个位偏差
                tongji_3 += 1  # 统计出现此卦的总数
                if ws_gailvfenxi[j][6].value == 0:  # 统计命中的个数
                    tongji_30 += 1
                    tongji_31 += 1
                    tongji_32 += 1
                if ws_gailvfenxi[j][6].value == 1:  # 统计命中的个数
                    tongji_31 += 1
                    tongji_32 += 1
                if ws_gailvfenxi[j][6].value == 2:  # 统计命中的个数
                    tongji_32 += 1
        print('卦号：'+ str(i+1)
              +'百位'+str(tongji_10/tongji_1 if tongji_1!=0 else 0)+';'+str(tongji_11/tongji_1 if tongji_1!=0 else 0)+';'+str(tongji_12/tongji_1 if tongji_1!=0 else 0)
              +'十位'+str(tongji_20/tongji_2 if tongji_2!=0 else 0)+';'+str(tongji_21/tongji_2 if tongji_2!=0 else 0)+';'+str(tongji_22/tongji_2 if tongji_2!=0 else 0)
              +'个位'+str(tongji_30/tongji_3 if tongji_3!=0 else 0)+';'+str(tongji_31/tongji_3 if tongji_3!=0 else 0)+';'+str(tongji_32/tongji_3 if tongji_3!=0 else 0))
        arr1.append(tongji_11 / tongji_1 if tongji_1!=0 else 0)
        arr2.append(tongji_21 / tongji_2 if tongji_2!=0 else 0)
        arr3.append(tongji_31 / tongji_3 if tongji_3!=0 else 0)
    print('\n# 偏差+-1直选')
    print('\n# 单卦单位准确率')
    print('\n# 偏差+-1直选盈亏')
    print('\n# 偏差+-1组选')
    print('\n# 单卦单位准确率')
    print('\n# 单卦单位汇总准确率')
    avg = 0
    for i in arr1:
        avg += i
    avg /= 64
    print(str(avg))
    print('\n# 偏差+-1组选盈亏')
    print('\n# 偏差+-2组选')
    print('\n# 单卦单位准确率')
    print('\n# 偏差+-2组选盈亏')


def analyse_1(typ, path1, path2, path3):
    realpath1 = os.path.join(os.getcwd(), path1)
    realpath2 = os.path.join(os.getcwd(), path2)
    realpath3 = os.path.join(os.getcwd(), path3)
    # 读取表格
    wb_geshihua = openpyxl.load_workbook(realpath1)
    ws_geshihua = wb_geshihua.active
    wb_huiguimoxing = openpyxl.load_workbook(realpath2)
    ws_huiguimoxing = wb_huiguimoxing.active
    if typ == "福彩3D":
        print('# 福彩3D：')
        # 概率分析表格初始化
        wb_gailvfenxi, ws_gailvfenxi = fucai3d_1(realpath3)
        for i, row in enumerate(ws_geshihua):
            # 每一个开奖数据都要判断，每一条数据三位三卦匹配结果输出到概率分析表格中供后面使用
            arr = fucai3d_2(ws_geshihua, ws_huiguimoxing, i)
            print(arr)
            if arr == 0:
                pass
            else:
                ws_gailvfenxi.append(arr)
        wb_gailvfenxi.save(realpath3)
    elif typ == "福彩双色球":
        pass
    else:
        print("无")


def analyse_2(typ, path3):
    realpath3 = os.path.join(os.getcwd(), path3)
    if typ == "福彩3D":
        fucai3d_3(realpath3)
    elif typ == "福彩双色球":
        pass
    else:
        print("无")
